import openpyxl
from datetime import datetime
import os

# Load the existing workbook
eval_path = f"{datetime.today().strftime('%Y%m%d')}\\"
filepaths = ["ragflow-hr_rpt.xlsx", "ragflow-hr_ocr.xlsx"]
# filepaths = os.listdir(eval_path)


def format_xlsx(filepath):
    workbook = openpyxl.load_workbook(filepath)

    # Dataset (the name in Json) |  Document | Query | ref_answer | Returned Docs | Context chunks | Response | Comment | Score
    target_cols_width = [15, 30, 30, 30, 30, 40, 30, 30, 10]

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # all cells wrap text and center align
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

        # set row height of the 2nd+ rows to 180
        sheet.row_dimensions[1].height = 15
        for i in range(2, sheet.max_row+1):
            sheet.row_dimensions[i].height = 180

        # set column width
        for i, width in enumerate(target_cols_width):
            sheet.column_dimensions[chr(65+i)].width = width

    # Save the modified workbook back to the file
    workbook.save(filepath)

if __name__ == "__main__":
    for file in filepaths:
        format_xlsx(eval_path + file)